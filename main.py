from openpyxl.reader.excel import load_workbook
import datetime
import matplotlib.pyplot as plt
import pandas as pd

wb = load_workbook("test.xlsx")
ws = wb.active


# Fukncja dodająca rekordydo pliku .xlsx
def dodaj():
    kwota = input("Podaj kwote: ")
    dzien = input("Dzień: ")
    if int(dzien) <= 0 or int(dzien) >= 32:
        print("Zły dzień")
        return
    miesiac = input("Miesiąc: ")
    if int(miesiac) <= 0 or int(miesiac) >= 13:
        print("Zły miesiac")
        return
    rok = input("Rok: ")
    if int(rok) <= 0:
        print("Rok nie może być ujemy!")
        return
    kategoria = input("Kategoria: ")
    opis = input("Opis: ")

    ws["A" + str(2 + int(ws["E2"].value))] = int(kwota)
    ws["B" + str(2 + int(ws["E2"].value))] = str(dzien + "." + miesiac + "." + rok)
    ws["C" + str(2 + int(ws["E2"].value))] = str(kategoria)
    ws["D" + str(2 + int(ws["E2"].value))] = str(opis)

    ws["E2"] = int(ws["E2"].value) + 1
    wb.save("test.xlsx")


# Fukncja wyświetlająca nagłówki kolumn
def naglowki():
    print(
        str(ws["A1"].value) + "\t" + str(ws["B1"].value) + "\t\t\t" + str(ws["C1"].value) + "\t" + str(ws["D1"].value))


# Fukncja wyświetlająca wartości w podanym wierszu
def wyswietl(x):
    print(str(ws["A" + str(x)].value) + ";\t" + str(ws["B" + str(x)].value) + ";\t\t" +
          str(ws["C" + str(x)].value) + ";\t\t" + str(ws["D" + str(x)].value))


# Funkcja, która czyści całą bazę
def wyczysc_wszystko():
    for x in range(2, 2 + int(ws["E2"].value)):
        ws.delete_rows(2)

    ws["E2"] = 0
    wb.save("test.xlsx")


# Fukncja szukająca wpisu po zadanej dacie, miesiącu lub w danym roku
def szukaj_data():
    dzien = input("Dzień: ")
    if int(dzien) <= 0 or int(dzien) >= 32:
        print("Zły dzień")
        return
    miesiac = input("Miesiąc: ")
    if int(miesiac) <= 0 or int(miesiac) >= 13:
        print("Zły miesiac")
        return
    rok = input("Rok: ")
    if int(rok) <= 0:
        print("Rok nie może być ujemy!")
        return
    naglowki()
    for x in range(2, 2 + int(ws["E2"].value)):
        if dzien == "" and miesiac == "" and rok == str(ws["B" + str(x)].value)[6:10]:
            wyswietl(x)
        elif dzien == "" and miesiac == str(ws["B" + str(x)].value)[3:5] and rok == str(ws["B" + str(x)].value)[6:10]:
            wyswietl(x)
        elif dzien == str(ws["B" + str(x)].value)[0:2] and miesiac == str(ws["B" + str(x)].value)[3:5] and rok == str(
                ws["B" + str(x)].value)[6:10]:
            wyswietl(x)


# Fukncja szukająca wpisu po zadanym okresie
def okres():
    dzien_p = int(input("Dzień początkowy: "))
    miesiac_p = int(input("Miesiąc początkowy: "))
    rok_p = int(input("Rok początkowy: "))
    data_p = datetime.date(rok_p, miesiac_p, dzien_p)

    dzien_k = int(input("Dzień końcowy: "))
    miesiac_k = int(input("Miesiąc końcowy: "))
    rok_k = int(input("Rok końcowy: "))
    data_k = datetime.date(rok_k, miesiac_k, dzien_k)

    naglowki()
    for x in range(2, 2 + int(ws["E2"].value)):
        data_x = datetime.date(int(str(ws["B" + str(x)].value)[6:10]), int(str(ws["B" + str(x)].value)[3:5]),
                               int(str(ws["B" + str(x)].value)[0:2]))
        if data_p <= data_x <= data_k:
            wyswietl(x)


# Fukncja szukająca wpisu po zadanej kategorii
def szukaj_kategoria():
    kat = []
    print("Aktualne dostępne kategorie:")
    for k in range(2, 2 + int(ws["E2"].value)):
        if str(ws["C" + str(k)].value) not in kat:
            kat.append(str(ws["C" + str(k)].value))
            print(str(ws["C" + str(k)].value))

    kategoria = input("\nKaregoria do wyszukania: ")
    naglowki()
    for x in range(2, 2 + int(ws["E2"].value)):
        if kategoria == str(ws["C" + str(x)].value):
            wyswietl(x)


# Funckaj rusyjąca wykresy i statystyki
def wykres():
    x = []
    y = []

    wyk = input("Wykres ma być:\n"
                "1. Liniowy\n"
                "2. Słupkowy\n"
                "(Dowolna wartość) Powrót\n")
    if wyk != "1" and wyk != "2":
        return

    dzien_p = int(input("Dzień początkowy: "))
    miesiac_p = int(input("Miesiąc początkowy: "))
    rok_p = int(input("Rok początkowy: "))
    data_p = datetime.date(rok_p, miesiac_p, dzien_p)

    dzien_k = int(input("Dzień końcowy: "))
    miesiac_k = int(input("Miesiąc końcowy: "))
    rok_k = int(input("Rok końcowy: "))
    data_k = datetime.date(rok_k, miesiac_k, dzien_k)

    p = pd.read_excel("test.xlsx")

    for n in range(2, 2 + int(ws["E2"].value)):
        data_x = datetime.date(int(str(ws["B" + str(n)].value)[6:10]), int(str(ws["B" + str(n)].value)[3:5]),
                               int(str(ws["B" + str(n)].value)[0:2]))
        if data_p <= data_x <= data_k:
            x.append(str(ws["B" + str(n)].value))
            y.append(int(ws["A" + str(n)].value))
    if wyk == "1":
        plt.plot(x, y)  # Wykres liniowy
    elif wyk == "2":
        plt.bar(x, y)  # Wykres słupkowy

    plt.xlabel("Data")
    plt.ylabel("Kwota")
    p_srednia = p["Kwota"].mean()
    plt.axhline(p_srednia, color="red")
    p_mediana = p["Kwota"].median()
    plt.axhline(p_mediana, color="green")
    p_dominanta = p.mode().iloc[0, 0]
    plt.axhline(p_dominanta, color="pink")
    plt.show()
    print("Statystiki tekstowo:")
    print(p.describe())

#Odpowiada za działanie menu
if __name__ == '__main__':
    while True:
        n = input("\nCo chcesz zrobić?\n"
                  "1. Dodaj wydatek\n"
                  "2. Wyświetl wszystkie wydatki\n"
                  "3. Szukaj\n"
                  "4. Wykres \n"
                  "9. Wyczyść cały kalkulator\n"
                  "0. Zakończ program\n")
        if n == "1":
            dodaj()
        elif n == "2":
            naglowki()
            for x in range(2, 2 + int(ws["E2"].value)):
                wyswietl(x)
            print("\nSuma wszystkich wydatków: " + str(ws["B1"].value))
        elif n == "3":
            print("Chcę szukać po:")
            s = input("1. Dacie\n"
                      "2. Okresie\n"
                      "3. Kategorii\n"
                      "(Dowolna wartość) Wróć\n")
            if s == "1":
                szukaj_data()
            elif s == "2":
                okres()
            elif s == "3":
                szukaj_kategoria()
            else:
                continue

        elif n == "4":
            wykres()
        elif n == "9":
            print("Na pewno chcesz wyczyścić całą historię? Sracisz w ten sposób wszyskie dane")
            w = input("(Dowolna wartość) Nie! Nie chcę stracić moich cennych danych!\n"
                      "9. Tak, chcę wyczyścić całą historię i zacząć od zera\n")
            if w == "9":
                wyczysc_wszystko()
            else:
                continue

        elif n == "0":
            print("\nPaPa...\n")
            exit()

        else:
            print("Podano złą wartość!\n")
    wb.save("test.xlsx")
